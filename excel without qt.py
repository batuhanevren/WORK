# -*- coding: utf-8 -*-
"""
Created on Mon Jan  1 10:54:36 2018

@author: batuhan
"""

import openpyxl
def total_excel(file_name):
    file_name = input('Please write the file name: ')
    
    data = openpyxl.load_workbook(file_name)

    data_sheet = data.active
    rows = data_sheet.max_row
    columns = data_sheet.max_column

    a = {}
    b = {}
    fee1 = list()
    fee2 = list()
    total = list()

    for r in range(1,rows+1):
        for c in range(1,columns+1):
            e = data_sheet.cell(row=r,column=1)
            a.update({e.value:0})
    for q in a.keys():
        fee1.append(q)
        
        
    for r in range(1,rows+1):
        for c in range(1,columns+1):
            e = data_sheet.cell(row=r,column=2)
            b.update({e.value:0})
        
    for q in b.keys():
        fee2.append(q)

    for i in range (len(fee1)):  
        total.append(fee1[i] + fee2[i])
    
   

    z = 0 
    for r in range(1,rows+1):
        if z <= len(total):
            t=data_sheet.cell(row=r,column=3)
            z += 1
        t.value = total[z-1]
    
    t=data_sheet.cell(row=1,column=3,value='TOTAL')
    
    data.save('Employees Datas2.xlsx')