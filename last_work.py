
import sys
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QTableWidgetItem, QTableWidget, QApplication
import openpyxl


class Ui_FeeCalculation(QtWidgets.QMainWindow):
    def setupUi(self, FeeCalculation):
        FeeCalculation.setObjectName("FeeCalculation")
        FeeCalculation.resize(777, 600)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap("todachoppa.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        FeeCalculation.setWindowIcon(icon)
        self.centralwidget = QtWidgets.QWidget(FeeCalculation)
        self.centralwidget.setObjectName("centralwidget")
        self.feeTable = QtWidgets.QTableWidget(self.centralwidget)
        self.feeTable.setGeometry(QtCore.QRect(0, 0, 550, 200))
        self.feeTable.viewport().setProperty("cursor", QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.feeTable.setMouseTracking(False)
        self.feeTable.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.feeTable.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.feeTable.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.feeTable.setTabKeyNavigation(True)
        self.feeTable.setShowGrid(True)
        self.feeTable.setRowCount(5)
        self.feeTable.setColumnCount(5)
        self.feeTable.setObjectName("feeTable")
        item = QtWidgets.QTableWidgetItem()
        item.setTextAlignment(QtCore.Qt.AlignCenter)
        font = QtGui.QFont()
        font.setBold(False)
        font.setWeight(50)
        item.setFont(font)
        self.feeTable.setItem(0, 0, item)
        self.values = QtWidgets.QPushButton(self.centralwidget)
        self.values.setGeometry(QtCore.QRect(600, 50, 75, 23))
        self.values.setObjectName("values")
        #excelden değerleri getirmek için atandı
        self.values.clicked.connect(self.excel)
        
        
        self.calculate = QtWidgets.QPushButton(self.centralwidget)
        self.calculate.setGeometry(QtCore.QRect(600, 90, 75, 23))
        self.calculate.setObjectName("calculate")
        # 'calculate' butonu için hesaplama atandı
        self.calculate.clicked.connect(self.calculation)
        
        
        self.quit = QtWidgets.QPushButton(self.centralwidget)
        self.quit.setGeometry(QtCore.QRect(600, 130, 75, 23))
        self.quit.setObjectName("quit")
        #'quit' butonu için için çıkış atandı.
        self.quit.clicked.connect(self.close_application)
        
        
        
        FeeCalculation.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(FeeCalculation)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 777, 21))
        self.menubar.setObjectName("menubar")
        self.mainMenu = QtWidgets.QMenu(self.menubar)
        self.mainMenu.setObjectName("mainMenu")
        FeeCalculation.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(FeeCalculation)
        self.statusbar.setObjectName("statusbar")
        FeeCalculation.setStatusBar(self.statusbar)
        self.actionOpen = QtWidgets.QAction(FeeCalculation)
        self.actionOpen.setObjectName("actionOpen")
        self.actionSave = QtWidgets.QAction(FeeCalculation)
        self.actionSave.setObjectName("actionSave")
        self.actionExit = QtWidgets.QAction(FeeCalculation)
        self.actionExit.setObjectName("actionExit")
        #menüdeki 'exit' için çıkış atandı.
        self.actionExit.triggered.connect(self.close_application)
        
        self.mainMenu.addAction(self.actionOpen)
        self.mainMenu.addAction(self.actionSave)
        self.mainMenu.addSeparator()
        self.mainMenu.addAction(self.actionExit)
        self.menubar.addAction(self.mainMenu.menuAction())

        self.retranslateUi(FeeCalculation)
        QtCore.QMetaObject.connectSlotsByName(FeeCalculation)

    def retranslateUi(self, FeeCalculation):
        _translate = QtCore.QCoreApplication.translate
        FeeCalculation.setWindowTitle(_translate("FeeCalculation", "MainWindow"))
        self.feeTable.setSortingEnabled(False)
        __sortingEnabled = self.feeTable.isSortingEnabled()
        self.feeTable.setSortingEnabled(False)
        item = self.feeTable.item(0, 0)
        item.setText(_translate("FeeCalculation", "5"))
        self.feeTable.setSortingEnabled(__sortingEnabled)
        self.values.setText(_translate("FeeCalculation", "Values"))
        self.calculate.setText(_translate("FeeCalculation", "Calculate"))
        self.quit.setText(_translate("FeeCalculation", "Quit"))
        
        
        
        
        self.mainMenu.setTitle(_translate("FeeCalculation", "Menu"))
        self.actionOpen.setText(_translate("FeeCalculation", "Open"))
        self.actionSave.setText(_translate("FeeCalculation", "Save"))
        self.actionExit.setText(_translate("FeeCalculation", "Exit"))
        
    #çıkış fonksiyonu tanımlandı.    
    def close_application(self):
        selection = QtWidgets.QMessageBox.question(self, 'Exit!',"Do you want to exit?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No )
        
        if selection == QtWidgets.QMessageBox.Yes:
            sys.exit()
        else:
            pass

    #Excelden değerleri getirmek üzere tanımlandı.           
    def excel(self):
        data = openpyxl.load_workbook('Employees Datas.xlsx')
        data_sheet = data.active
        rows = data_sheet.max_row
        columns = data_sheet.max_column
        self.feeTable.setRowCount(rows)
        self.feeTable.setColumnCount(columns)       
        for r in range(1,rows+1):
            for c in range(1,columns+1):
                x = data_sheet.cell(row=r,column=c).value
                item = QTableWidgetItem(str(x))
                self.feeTable.setItem(r-1,c-1,item) 
                
                
    #Tablodaki değerler üzeinden işlem yapılmak üzere tanımlandı.    
    def calculation(self):                             
        data = openpyxl.load_workbook('Employees Datas.xlsx')
        data_sheet = data.active
        rows = data_sheet.max_row
        columns = data_sheet.max_column                      
        for r in range(0,rows+1):  
            x = self.feeTable.item(r+1,0)
            x = int(x.text()) 
            y = self.feeTable.item(r+1,1)
            y = int(y.text()) 
            t = x + y
            item = QTableWidgetItem(str(t))
            self.feeTable.setItem(r+1,2,item)
          
        
        
        
        
if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    FeeCalculation = QtWidgets.QMainWindow()
    ui = Ui_FeeCalculation()
    ui.setupUi(FeeCalculation)
    FeeCalculation.show()
    sys.exit(app.exec_())







