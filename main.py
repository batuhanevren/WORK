# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'main.ui'
#
# Created by: PyQt5 UI code generator 5.6
#
# WARNING! All changes made in this file will be lost!

from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3
import openpyxl

class Ui_MainWindow(object):
    def calculate(self):
        
        data = openpyxl.load_workbook('Employees Datas.xlsx')
        data_sheet = data.active
        rows = data_sheet.max_row
        columns = data_sheet.max_column
        a = {}
        b = {}
        fee1 = list()
        fee2 = list()
        total = list()
        for r in range(1,rows+1):
            for c in range(1,columns+1):
                e = data_sheet.cell(row=r,column=1)
                a.update({e.value:0})
        for q in a.keys():
                fee1.append(q)            
        for r in range(1,rows+1):
            for c in range(1,columns+1):
                e = data_sheet.cell(row=r,column=2)
                b.update({e.value:0})       
        for q in b.keys():
            fee2.append(q)
        for i in range (len(fee1)):  
            total.append(fee1[i] + fee2[i])      
        z = 0 
        for r in range(1,rows+1):
            if z <= len(total):
                t=data_sheet.cell(row=r,column=3)
                z += 1
            t.value = total[z-1]    
        t=data_sheet.cell(row=1,column=3,value='TOTAL')    
        data.save('Employees Datas2.xlsx')
    
    
    
    def loadData(self):
        connection = sqlite3.connect('my_db.db')
        query = "SELECT * FROM my_db"
        result = connection.execute(query)
        self.tableWidget.setRowCount(0)
        
        for row_number, row_data in enumerate(result):
            self.tableWidget.insertRow(row_number)
            for column_number, data in enumerate(row_data):
                self.tableWidget.setItem(row_number, column_number, QtWidgets.QTableWidgetItem(str(data)))
        connection.close()
        
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(330, 513)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(0, 0, 321, 211))
        self.tableWidget.setRowCount(20)
        self.tableWidget.setColumnCount(20)
        self.tableWidget.setObjectName("tableWidget")
        self.btn_load = QtWidgets.QPushButton(self.centralwidget)
        self.btn_load.setGeometry(QtCore.QRect(130, 350, 75, 23))
        self.btn_load.setObjectName("btn_load")
        self.btn_load.clicked.connect(self.calculate)
        
        
        
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 330, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.btn_load.setText(_translate("MainWindow", "Load"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())

