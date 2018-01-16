# -*- coding: utf-8 -*-
"""
Created on Mon Jan  1 10:54:36 2018

@author: batuhan
"""

import openpyxl
   
data = openpyxl.load_workbook('Employees Datas.xlsx')

data_sheet = data.active
rows = data_sheet.max_row
columns = data_sheet.max_column
for r in range(1,rows+1):
    for c in range(1,columns+1):
        x = data_sheet.cell(row=r,column=c).value
        print (x)